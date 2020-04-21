#-*-coding: utf-8 -*-
from bs4 import BeautifulSoup
import time, certifi, ssl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
options = Options()
options.headless = False#True
options.add_argument('--ignore-certificate-errors')
'''
options.add_argument('--ignore-certificate-errors')
mobile_emulation = {"deviceMetrics": { "width": 360, "height": 640, "pixelRatio": 3.0 },"userAgent": "Mozilla/5.0 (Linux; Android 4.2.1; en-us; Nexus 5 Build/JOP40D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166 Mobile Safari/535.19" }
options.add_experimental_option("mobileEmulation", mobile_emulation)
'''
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "No."
ws.cell(row=1, column=2).value = "URL"
ws.cell(row=1, column=3).value = "data"
driver = webdriver.Chrome(executable_path='C:\driver\chromedriver.exe',options=options)
http_protocol="http"
https_protocol="https"
def readList():
    file=open('C:\\Users\\9000784\\Downloads\\asd.txt','r',encoding='UTF8')
    All=file.read()
    print(All)
    file.close()
    lines=All.split('\n')
    print(lines)
    return accessURL(lines)

def accessURL(index):
    try:
        with open('C:\\driver\\targets_url.txt', 'r') as f:
            bigdata=f.read()
            bigdata=bigdata.split('\n')
            for s in bigdata:
                index=index+1
                time.sleep(1)
                driver.implicitly_wait(2)
                driver.get('{}://{}'.format(http_protocol,s))
                #driver = webdriver.Remote(command_executor='{}://{}'.format(http_protocol,s),desired_capabilities = chrome_options.to_capabilities())
                html = driver.page_source # 페이지의 elements모두 가져오기
                print("총 {}중 {}번 완료".format('195',index))
                #print(s, end='')
                #soup = BeautifulSoup(html, 'html.parser') # BeautifulSoup사용하기
                #soup.prettify()
                driver.save_screenshot('C:\\driver\\capture\\{}-{}-{}.png'.format(index,http_protocol,s))
                writeResult(index, bigdata, html)
                #driver.close()
    except Exception as e:
        print("[+]에러 : {}".format(e))
            
def writeResult(index, lines, data):
    ws.cell(row=index+1, column=2).value = lines[(index-1)]
    ws.cell(row=index+1, column=1).value = index
    ws.cell(row=index+1, column=3).value = data

if __name__ == "__main__":
    index=0
    accessURL(index)
    wb.save('C:\\driver\\check_service_{}.xlsx'.format(http_protocol))
    
    
'''   
        except HTTPError as e:
            code = e.getcode()
            writeResult(index, lines, code)
        except URLError as e:
            writeResult(index, lines, "URLError")
'''
