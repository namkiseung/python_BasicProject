#-*-coding: utf-8 -*-
import time, certifi, ssl
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from urllib.parse import urlparse, urlencode
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "No."
ws.cell(row=1, column=2).value = "URL"
ws.cell(row=1, column=3).value = "Status"

def accessURL(index):
    with open('C:\\driver\\targets_url.txt', 'r') as f:
        bigdata=f.read()
        bigdata=bigdata.split('\n')
        for s in bigdata:
            index=index+1
            print("총 {}중 {}번 완료".format('195',index))
            try:  
                context = ssl._create_unverified_context()
                https_res = urlopen('https://{}'.format(s), context=context)
                http_status = https_res.status
                result_success="443 port : {}".format(http_status)
                print(result_success)
                writeResult(index, bigdata, result_success)
            except HTTPError as e:
                code = e.getcode()
                result_success="443 port : {}".format(code)
                print(result_success)
                writeResult(index, bigdata, result_success)
            except URLError as e:
                result_success="443 port : {}".format(e)
                print(result_success)
                writeResult(index, bigdata, result_success)
            except Exception as e:
                result_success="443 port : {}".format(e)
                print(result_success)
                writeResult(index, bigdata, result_success)              
   
def writeResult(index, lines, data):
    ws.cell(row=index+1, column=2).value = lines[(index-1)]
    ws.cell(row=index+1, column=1).value = index
    ws.cell(row=index+1, column=3).value = data

if __name__ == "__main__":
    index=0
    accessURL(index)
    wb.save('C:\\driver\\Check_status_https_0420.xlsx')