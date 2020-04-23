import dns.resolver
import re
import certifi
import ssl
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "No."
ws.cell(row=1, column=2).value = "URL"
ws.cell(row=1, column=3).value = "Status(HTTPS)"
ws.cell(row=1, column=4).value = "Status(HTTP)"
ws.cell(row=1, column=5).value = "IP List"

rt_a="IN A"
rt_aa="IN AA"
rt_cn="IN CNAME"

def readList():
    file=open('C:\\Users\\9000784\\Downloads\\domain2.txt','r',encoding='UTF8')
    All=file.read()
    file.close()
    lines=All.split('\n')
    return accessURL(lines)

def formating_text(mention):
    mention="{}".format(mention.replace(rt_a,""))
    mention=mention.replace('\n','')
    mention=mention.replace('. ',' | ')
    mention=re.findall('[0-9]+(?:\.[0-9]+){3}', mention)
    print(mention)
    return mention

def accessURL(lines):
    for index, value in enumerate(lines, start=1):
        a = lines[index-1].split('/')
        strLines = str(a[0])
        domain = strLines.strip('') #strip작업
        r = dns.resolver.Resolver()#dns 인터페이스 사용
        r.nameservers = ['8.8.8.8'] #nameserver지정
        ip_list = []
        try:
            #print(dns.resolver.query(domain).response.answer) #log 0x01
            for answer in dns.resolver.query(domain).response.answer:
                idx_ip="{}".format(answer)
                #print(idx_ip) #log 0x02
                #만약 IN A가 존재한다면
                if int(idx_ip.find(rt_a)) != -1:
                    #개수만큼 반복해서 정규화 거치고 append하자
                    for trans in formating_text(idx_ip):
                        #print('trans len :', len(trans)) #log 0x03
                        ip_list.append("{}".format(trans))
                #만약 IN AA가 존재한다면
                if int(idx_ip.find(rt_aa)) != -1:
                    #개수만큼 반복해서 정규화 거치고 append하자
                    for trans in formating_text(idx_ip):
                        ip_list.append("{}".format(trans))
                #만약 IN CNAME 존재한다면
                if int(idx_ip.find(rt_cn)) != -1:                        
                    #개수만큼 반복해서 정규화 거치고 append하자
                    for trans in formating_text(idx_ip):
                        ip_list.append("{}".format(trans))
        except Exception as ex:
            if 'None of DNS query names exist' in str(ex):
                message = 'None'
            else:
                message = 'Error!'
        ip_list = str(ip_list).encode()+str("\n").encode()
        
        try:
            context = ssl._create_unverified_context()
            res = urlopen('https://'+value, context=context)
            status = res.status
            
        except HTTPError as e:
            status = e.getcode()
            
        except URLError as e:
            status = "URLError"
            
##############
        try:
            context = ssl._create_unverified_context()
            res2 = urlopen('http://'+value, context=context)
            status2 = res2.status
            
        except HTTPError as e:
            status2 = e.getcode()
            
        except URLError as e:
            status2 = "URLError"
        writeResult(index, lines, status, status2, ip_list)    

def writeResult(index, lines, status, status2, ip_list):
    ws.cell(row=index+1, column=1).value = index
    ws.cell(row=index+1, column=2).value = lines[(index-1)]
    ws.cell(row=index+1, column=3).value = status
    ws.cell(row=index+1, column=4).value = status2
    ws.cell(row=index+1, column=5).value = ip_list


if __name__ == "__main__":
    readList()
    wb.save('C:\\Users\\9000784\\Downloads\\test222.xlsx')