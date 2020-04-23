#-*-coding: utf-8 -*-
from bs4 import BeautifulSoup
import time, certifi, ssl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
options = Options()
options.headless = False#True
options.add_argument('--ignore-certificate-errors')
#mobile_emulation = {"deviceMetrics": { "width": 360, "height": 640, "pixelRatio": 3.0 },"userAgent": "Mozilla/5.0 (Linux; Android 4.2.1; en-us; Nexus 5 Build/JOP40D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166 Mobile Safari/535.19" }
#options.add_experimental_option("mobileEmulation", mobile_emulation)

wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "No."
ws.cell(row=1, column=2).value = "URL"
ws.cell(row=1, column=3).value = "data"
driver = webdriver.Chrome(executable_path='C:\driver\chromedriver.exe',options=options)
http_protocol="http"
https_protocol="https"
def readList():
    file=open('C:\\Users\\9000784\\Downloads\\asd.txt','r',encoding='UTF8')
    All=file.read()
    print(All)
    file.close()
    lines=All.split('\n')
    print(lines)
    return accessURL(lines)

def accessURL(index):
    try:
        with open('C:\\driver\\targets_url.txt', 'r') as f:
            bigdata=f.read()
            bigdata=bigdata.split('\n')
            for s in bigdata:
                index=index+1
                time.sleep(1)
                driver.implicitly_wait(2)
                driver.get('{}://{}'.format(https_protocol,s))
                #driver = webdriver.Remote(command_executor='{}://{}'.format(http_protocol,s),desired_capabilities = chrome_options.to_capabilities())
                html = driver.page_source # 페이지의 elements모두 가져오기
                print("총 {}중 {}번 완료".format('195',index))
                #print(s, end='')
                #soup = BeautifulSoup(html, 'html.parser') # BeautifulSoup사용하기
                #soup.prettify()
                driver.save_screenshot('C:\\driver\\capture\\{}-{}-{}.png'.format(index,https_protocol,s))
                writeResult(index, bigdata, html)
                #driver.close()
    except Exception as e:
        print("[+]에러 : {}".format(e))
            
def writeResult(index, lines, data):
    ws.cell(row=index+1, column=2).value = lines[(index-1)]
    ws.cell(row=index+1, column=1).value = index
    ws.cell(row=index+1, column=3).value = data

if __name__ == "__main__":
    index=0
    accessURL(index)
    wb.save('C:\\driver\\check_service_{}.xlsx'.format(https_protocol))
    
    
'''   
        except HTTPError as e:
            code = e.getcode()
            writeResult(index, lines, code)
        except URLError as e:
            writeResult(index, lines, "URLError")
>>reference
1. https://chromedriver.chromium.org/mobile-emulation
2. https://hanmaruj.tistory.com/3
3. 에러 https://korbillgates.tistory.com/91
4. 스크립트 삽입 https://hashcode.co.kr/questions/8032/%ED%8C%8C%EC%9D%B4%EC%8D%AC-%EC%85%80%EB%A0%88%EB%8B%88%EC%9B%80%EC%97%90%EC%84%9C-webdriverclose%EB%A1%9C-%ED%81%AC%EB%A1%AC-%EC%B0%BD%EC%9D%84-%EB%8B%AB%EC%9D%80-%ED%9B%84%EC%97%90-%EB%8B%A4%EC%8B%9C-%ED%81%AC%EB%A1%AC%EC%B0%BD%EC%9D%84-%EC%83%9D%EC%84%B1%EC%8B%9C%ED%82%A4%EB%8A%94-%EB%B0%A9%EB%B2%95%EC%A7%88%EB%AC%B8
5. ssl https://cnpnote.tistory.com/entry/PYTHON-Selenium%EC%9D%84-%EC%82%AC%EC%9A%A9%ED%95%98%EC%97%AC-%EC%9D%B8%EC%A6%9D%EC%84%9C%EB%A5%BC-%EC%B2%98%EB%A6%AC%ED%95%98%EB%8A%94-%EB%B0%A9%EB%B2%95%EC%9D%80-%EB%AC%B4%EC%97%87%EC%9E%85%EB%8B%88%EA%B9%8C
6. ssl에러 https://testmanager.tistory.com/147
7. 캡쳐 https://marshawn.tistory.com/121
8. 캡처 https://hashcode.co.kr/questions/9490/%ED%8C%8C%EC%9D%B4%EC%8D%AC-selenium%EC%9C%BC%EB%A1%9C-%ED%8E%98%EC%9D%B4%EC%A7%80-%EC%A0%84%EC%B2%B4-%EC%8A%A4%ED%81%AC%EB%A6%B0%EC%83%B7-%EB%B0%A9%EB%B2%95
9. 캡처 https://studyforus.com/share/584148
10 시작 https://beomi.github.io/2017/02/27/HowToMakeWebCrawler-With-Selenium/

'''